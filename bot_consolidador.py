import pandas as pd
import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from datetime import datetime

class ExcelBotGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Bot Excel")
        self.root.geometry("800x600")
        self.root.configure(bg="white")

        self.origem_path = ""
        self.destino_path = ""
        self.df_origem = None
        self.aba_origem = ""
        self.aba_destino = ""

        self.log_text = None
        self.criar_interface()

    def criar_interface(self):
        frame = ttk.Frame(self.root)
        frame.pack(pady=10)

        ttk.Button(frame, text="Selecionar Arquivo de Origem", command=self.selecionar_origem).grid(row=0, column=0, padx=5)
        ttk.Button(frame, text="Selecionar Arquivo de Destino", command=self.selecionar_destino).grid(row=0, column=1, padx=5)
        ttk.Button(frame, text="Iniciar Processo", command=self.iniciar_processo).grid(row=0, column=2, padx=5)

        self.progress = ttk.Progressbar(self.root, mode='indeterminate')
        self.progress.pack(pady=10, fill='x', padx=20)

        self.log_text = tk.Text(self.root, height=15)
        self.log_text.pack(fill='both', expand=True, padx=10, pady=10)

    def log(self, mensagem):
        self.log_text.insert('end', mensagem + "\n")
        self.log_text.see('end')
        self.root.update()

    def selecionar_origem(self):
        caminho = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xlsb *.xls *.csv")])
        if caminho:
            self.origem_path = caminho
            self.log(f"Origem selecionada: {caminho}")

    def selecionar_destino(self):
        caminho = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xlsb *.xls")])
        if caminho:
            self.destino_path = caminho
            self.log(f"Destino selecionado: {caminho}")

    def iniciar_processo(self):
        if not self.origem_path or not self.destino_path:
            messagebox.showerror("Erro", "Por favor, selecione ambos os arquivos.")
            return

        self.df_origem, self.aba_origem = self.ler_arquivo_com_aba(self.origem_path, "origem")
        if self.df_origem is None:
            return

        self.aplicar_filtros_e_inserir()

    def ler_arquivo_com_aba(self, caminho, tipo):
        try:
            if caminho.endswith('.csv'):
                df = pd.read_csv(caminho, encoding='utf-8', sep=None, engine='python')
                return df, None
            elif caminho.endswith('.xlsb'):
                xls = pd.ExcelFile(caminho, engine='pyxlsb')
            elif caminho.endswith('.xls'):
                xls = pd.ExcelFile(caminho, engine='xlrd')
            else:
                xls = pd.ExcelFile(caminho, engine='openpyxl')

            abas = xls.sheet_names
            aba_escolhida = self.selecionar_aba_popup(abas, tipo)
            if aba_escolhida is None:
                return None, None

            df = xls.parse(aba_escolhida)
            return df, aba_escolhida
        except Exception as e:
            messagebox.showerror("Erro ao ler o arquivo", str(e))
            return None, None

    def selecionar_aba_popup(self, abas, tipo):
        win = tk.Toplevel(self.root)
        win.title(f"Escolher aba do arquivo de {tipo}")
        ttk.Label(win, text=f"Escolha a aba do arquivo de {tipo}:").pack(pady=10)

        aba_var = tk.StringVar(value=abas[0])
        combo = ttk.Combobox(win, values=abas, textvariable=aba_var)
        combo.pack(pady=10)

        def confirmar():
            win.destroy()

        ttk.Button(win, text="Confirmar", command=confirmar).pack(pady=10)
        win.wait_window()
        return aba_var.get()

    def primeira_linha_vazia(self, ws, col_idx):
        linha = 1
        while ws.cell(row=linha, column=col_idx).value is not None:
            linha += 1
        return linha

    def ultima_linha_preenchida(self, ws, colunas):
        ultima_linha = 0
        for col_idx in colunas:
            linha = ws.max_row
            while linha > 0 and ws.cell(row=linha, column=col_idx).value is None:
                linha -= 1
            if linha > ultima_linha:
                ultima_linha = linha
        return ultima_linha + 1

    def aplicar_filtros_e_inserir(self):
        colunas = list(self.df_origem.columns)

        win = tk.Toplevel(self.root)
        win.title("Selecionar filtros e colunas de destino")
        win.geometry("900x700")

        ttk.Label(win, text="Para cada coluna, selecione valores a INCLUIR e a EXCLUIR e a coluna de destino (ex: B):").pack(pady=10)

        container = ttk.Frame(win)
        container.pack(fill='both', expand=True, padx=10, pady=5)

        canvas = tk.Canvas(container)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        filtro_widgets = []
        for col in colunas:
            frame = ttk.Frame(scrollable_frame)
            frame.pack(fill='x', padx=5, pady=3)

            ttk.Label(frame, text=col, width=20).grid(row=0, column=0, rowspan=2)

            valores_unicos = sorted(self.df_origem[col].dropna().astype(str).unique())
            var_inc = tk.Variable(value=valores_unicos)
            var_exc = tk.Variable(value=valores_unicos)

            lb_inc = tk.Listbox(frame, listvariable=var_inc, selectmode='multiple', height=4, exportselection=False)
            lb_exc = tk.Listbox(frame, listvariable=var_exc, selectmode='multiple', height=4, exportselection=False)

            lb_inc.grid(row=0, column=1, padx=5, sticky='ew')
            lb_exc.grid(row=1, column=1, padx=5, sticky='ew')

            ttk.Label(frame, text="Incluir").grid(row=0, column=2)
            ttk.Label(frame, text="Excluir").grid(row=1, column=2)

            col_entry = ttk.Entry(frame, width=5)
            col_entry.grid(row=0, column=3, rowspan=2, padx=5)

            btn_limpar_inc = ttk.Button(frame, text="Limpar Inc.", width=10, command=lambda lb=lb_inc: lb.selection_clear(0, 'end'))
            btn_limpar_inc.grid(row=0, column=4, padx=5)

            btn_limpar_exc = ttk.Button(frame, text="Limpar Exc.", width=10, command=lambda lb=lb_exc: lb.selection_clear(0, 'end'))
            btn_limpar_exc.grid(row=1, column=4, padx=5)

            filtro_widgets.append((col, lb_inc, lb_exc, col_entry))

        ttk.Label(win, text="Selecione colunas para remover duplicatas:").pack(pady=5)
        dup_var = tk.Variable(value=colunas)
        dup_listbox = tk.Listbox(win, listvariable=dup_var, selectmode='multiple', height=6, exportselection=False)
        dup_listbox.pack(fill='x', padx=20, pady=5)

        def confirmar():
            mapeamento = {}
            filtros_incluir = {}
            filtros_excluir = {}

            for col, lb_inc, lb_exc, entry in filtro_widgets:
                destino = entry.get().strip().upper()
                if not destino:
                    continue

                try:
                    destino_index = column_index_from_string(destino)
                except:
                    messagebox.showerror("Erro", f"Coluna de destino inválida para '{col}': {destino}")
                    return

                sel_inc = [lb_inc.get(i) for i in lb_inc.curselection()]
                sel_exc = [lb_exc.get(i) for i in lb_exc.curselection()]

                if sel_inc:
                    filtros_incluir[col] = sel_inc
                if sel_exc:
                    filtros_excluir[col] = sel_exc

                mapeamento[col] = destino_index

            if not mapeamento:
                messagebox.showerror("Erro", "Nenhuma coluna de destino definida.")
                return

            df_filtrado = self.df_origem.copy()

            for col, valores in filtros_incluir.items():
                df_filtrado = df_filtrado[df_filtrado[col].astype(str).isin(valores)]
            for col, valores in filtros_excluir.items():
                df_filtrado = df_filtrado[~df_filtrado[col].astype(str).isin(valores)]

            colunas_duplicadas = [dup_listbox.get(i) for i in dup_listbox.curselection()]
            if colunas_duplicadas:
                df_filtrado = df_filtrado.drop_duplicates(subset=colunas_duplicadas)

            if df_filtrado.empty:
                messagebox.showwarning("Aviso", "Nenhuma linha encontrada com os filtros aplicados.")
                return

            def exportar():
                if self.destino_path.endswith('.xlsb'):
                    messagebox.showerror("Erro", "Não é possível escrever em arquivos .xlsb. Use .xlsx.")
                    return
                elif self.destino_path.endswith('.xls'):
                    messagebox.showerror("Erro", "openpyxl não suporta escrita em .xls. Use .xlsx.")
                    return

                wb = load_workbook(self.destino_path)
                abas = wb.sheetnames
                aba_destino = self.selecionar_aba_popup(abas, "destino")
                ws = wb[aba_destino]

                colunas_indices = list(mapeamento.values())
                linha_inicial = self.ultima_linha_preenchida(ws, colunas_indices)

                for i, (_, row) in enumerate(df_filtrado.iterrows()):
                    for col_nome, col_dest_index in mapeamento.items():
                        valor = row[col_nome]
                        ws.cell(row=linha_inicial + i, column=col_dest_index, value=valor)

                pasta_saida = "exportados"
                os.makedirs(pasta_saida, exist_ok=True)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                nome_arquivo = f"resultado_{timestamp}.xlsx"
                caminho_final = os.path.join(pasta_saida, nome_arquivo) 

                try:
                    wb.save(caminho_final)
                except Exception as e:
                    messagebox.showerror("Erro ao salvar o arquivo", str(e))
                    return

                messagebox.showinfo("Sucesso", f"Dados inseridos em: {caminho_final}")
                preview_win.destroy()
                win.destroy()

            preview_win = tk.Toplevel(self.root)
            preview_win.title("Pré-visualização dos dados")
            preview_win.geometry("800x400")

            tree = ttk.Treeview(preview_win, columns=list(df_filtrado.columns), show='headings')
            for col in df_filtrado.columns:
                tree.heading(col, text=col)
                tree.column(col, width=100, anchor='w')
            for _, row in df_filtrado.iterrows():
                tree.insert('', 'end', values=list(row))
            tree.pack(fill='both', expand=True)

            frame_btns = ttk.Frame(preview_win)
            frame_btns.pack(pady=10)

            ttk.Button(frame_btns, text="Exportar para Excel", command=exportar).pack(side='left', padx=10)
            ttk.Button(frame_btns, text="Voltar e Editar Filtros", command=preview_win.destroy).pack(side='left', padx=10)

        ttk.Button(win, text="Confirmar e Visualizar", command=confirmar).pack(pady=10)

if __name__ == '__main__':
    root = tk.Tk()
    app = ExcelBotGUI(root)
    root.mainloop()